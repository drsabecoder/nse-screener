"""
NSE Stock Screener — Streamlit Web App
=======================================
Run with:
    pip install streamlit requests pandas openpyxl
    streamlit run nse_screener_app.py
"""

import streamlit as st
import requests
import pandas as pd
import time
import json
import os
import io
from datetime import datetime

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="NSE Stock Screener",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-title {font-size: 2rem; font-weight: 700; margin-bottom: 0;}
    .sub-title  {font-size: 1rem; color: #888; margin-top: 0;}
    .metric-card {
        background: #f8f9fa; border-radius: 12px;
        padding: 1rem 1.25rem; text-align: center;
        border: 1px solid #e9ecef;
    }
    .metric-val  {font-size: 2rem; font-weight: 700; margin: 0;}
    .metric-lbl  {font-size: 0.8rem; color: #888; margin: 0;}
    .valid-badge   {background:#d4edda; color:#155724; padding:3px 10px; border-radius:20px; font-size:12px; font-weight:600;}
    .watch-badge   {background:#fff3cd; color:#856404; padding:3px 10px; border-radius:20px; font-size:12px; font-weight:600;}
    .reject-badge  {background:#f8d7da; color:#721c24; padding:3px 10px; border-radius:20px; font-size:12px; font-weight:600;}
    .section-header {font-size:1.2rem; font-weight:600; margin: 1rem 0 0.5rem; border-bottom: 2px solid #e9ecef; padding-bottom: 6px;}
    div[data-testid="stButton"] button {width: 100%;}
    .stProgress > div > div {background: linear-gradient(90deg, #28a745, #20c997);}
</style>
""", unsafe_allow_html=True)

# ── Constants ──────────────────────────────────────────────────────────────────
BASE            = "https://nesliferesearch.com"
REQUEST_TIMEOUT = 30
CHECKPOINT_FILE = "nse_checkpoint.json"
MDSPORT_MC_MIN  = 1_000
MDSPORT_MC_MAX  = 10_000
MDSPORT_DE_MAX  = 0.30
MDSPORT_VALID   = 75
MDSPORT_WATCH   = 50
MPTDS_MC_MIN    = 10_000

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
    "Accept":     "application/json, text/plain, */*",
    "Referer":    f"{BASE}/education/",
    "Origin":     BASE,
}

NIFTY50 = [
    "RELIANCE","TCS","HDFCBANK","INFY","ICICIBANK","HINDUNILVR","ITC","SBIN",
    "BHARTIARTL","KOTAKBANK","LT","AXISBANK","ASIANPAINT","MARUTI","BAJFINANCE",
    "WIPRO","HCLTECH","ULTRACEMCO","POWERGRID","NTPC","SUNPHARMA","TITAN",
    "BAJAJFINSV","TECHM","NESTLEIND","TATAMOTORS","M&M","ADANIENT","ADANIPORTS",
    "COALINDIA","ONGC","JSWSTEEL","TATASTEEL","INDUSINDBK","HDFCLIFE","SBILIFE",
    "DRREDDY","CIPLA","DIVISLAB","EICHERMOT","APOLLOHOSP","BAJAJ-AUTO",
    "HEROMOTOCO","BRITANNIA","BPCL","TATACONSUM","GRASIM","HINDALCO","UPL","SHRIRAMFIN"
]

# ── Session state init ─────────────────────────────────────────────────────────
for key, val in {
    "running": False, "mds_results": {}, "mpt_results": {},
    "log": [], "total": 0, "done": 0
}.items():
    if key not in st.session_state:
        st.session_state[key] = val


# ══════════════════════════════════════════════════════════════════════════════
#  CORE SCREENER FUNCTIONS (same logic as nse_screener.py)
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=3600, show_spinner=False)
def fetch_nse_symbols():
    s = requests.Session()
    s.headers.update({"User-Agent": HEADERS["User-Agent"],
                       "Accept": "text/html,application/xhtml+xml"})
    try:
        s.get("https://www.nseindia.com", timeout=12)
        time.sleep(1.2)
        r = s.get("https://nsearchives.nseindia.com/content/equities/EQUITY_L.csv", timeout=30)
        r.raise_for_status()
        from io import StringIO
        df = pd.read_csv(StringIO(r.text))
        col = next(c for c in df.columns if "SYMBOL" in c.upper())
        return [x for x in df[col].dropna().astype(str).str.strip().tolist() if x and x != "nan"]
    except:
        pass
    try:
        s.get("https://www.nseindia.com", timeout=12); time.sleep(1)
        r = s.get("https://nsearchives.nseindia.com/content/indices/ind_nifty500list.csv", timeout=20)
        r.raise_for_status()
        from io import StringIO
        df = pd.read_csv(StringIO(r.text))
        col = next(c for c in df.columns if "Symbol" in c or "SYMBOL" in c.upper())
        return df[col].dropna().astype(str).str.strip().tolist()
    except:
        return NIFTY50


def score_param(value):
    if value < 5:  return 0
    if value < 10: return 1
    if value < 15: return 2
    if value < 20: return 3
    return 4


def safe(lst):
    out = []
    for x in lst:
        try: out.append(float(x) if x is not None else 0.0)
        except: out.append(0.0)
    return out[-6:] if len(out) >= 6 else out


def fetch_company_data(symbol, session):
    url = f"{BASE}/api/company/{symbol}"
    try:
        r = session.get(url, timeout=REQUEST_TIMEOUT)
        if r.status_code == 404: return {"_status": "NOT_FOUND"}
        if r.status_code == 429:
            time.sleep(6)
            r = session.get(url, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
        data = r.json(); data["_status"] = "OK"
        return data
    except requests.exceptions.Timeout:  return {"_status": "TIMEOUT"}
    except requests.exceptions.ConnectionError: return {"_status": "CONN_ERROR"}
    except Exception as e: return {"_status": "ERROR", "_error": str(e)}


def compute_mdsport(symbol, data):
    status = data.get("_status", "OK")
    if status != "OK":
        return {"symbol": symbol, "company_name": "", "market_cap_cr": None,
                "is_financial": False, "de_ratio": None, "de_pass": None,
                "mc_pass": False, "passes_gates": False, "total_score": None,
                "verdict": status, "s_score":0,"p_score":0,"o_score":0,"r_score":0,"t_score":0,
                "years_used": None, "status": status, "error": data.get("_error","")}

    fin    = data.get("financials", {})
    sales  = fin.get("sales", []);  ebitda = fin.get("ebitda", [])
    tax    = fin.get("tax", []);    pat    = fin.get("pat", [])
    te     = fin.get("te", []);     ncl    = fin.get("ncl", [])
    years  = data.get("years", [])
    mc     = float(data.get("marketCap") or 0)
    is_fin = bool(data.get("isFinancialCompany", False))
    name   = data.get("companyName", symbol)
    ticker = data.get("ticker", symbol)

    mc_pass  = MDSPORT_MC_MIN <= mc <= MDSPORT_MC_MAX
    de_ratio = None; de_pass = None
    if not is_fin and len(te) > 0 and len(ncl) > 0:
        te_last = te[-1]; ncl_last = ncl[-1]
        if te_last and te_last != 0:
            de_ratio = ncl_last / te_last
            de_pass  = de_ratio < MDSPORT_DE_MAX
        else:
            de_pass = False
    elif is_fin:
        de_pass = True

    passes_gates = mc_pass and (de_pass is True or de_pass is None)

    sales = safe(sales); ebitda = safe(ebitda); tax = safe(tax)
    pat   = safe(pat);   te     = safe(te);     ncl = safe(ncl)
    years = years[-5:]

    n = min(len(sales), len(ebitda), len(tax), len(pat), len(te), len(ncl))
    if n < 2:
        return {"symbol": ticker, "company_name": name, "market_cap_cr": mc,
                "is_financial": is_fin, "de_ratio": de_ratio, "de_pass": de_pass,
                "mc_pass": mc_pass, "passes_gates": passes_gates,
                "total_score": None, "verdict": "INSUFFICIENT_DATA",
                "s_score":0,"p_score":0,"o_score":0,"r_score":0,"t_score":0,
                "years_used": 0, "status": "INSUFFICIENT_DATA", "error": "< 2 years of data"}

    s_scores, p_scores, o_scores, r_scores, t_scores = [], [], [], [], []
    year_labels = []
    for i in range(1, n):
        s_val = ((sales[i]/sales[i-1])-1)*100 if sales[i-1] else 0.0
        p_val = ((pat[i]/pat[i-1])-1)*100     if pat[i-1]   else 0.0
        o_val = (ebitda[i]/sales[i])*100       if sales[i]   else 0.0
        cap   = ncl[i] + te[i]
        r_val = (pat[i]/cap)*100               if cap        else 0.0
        t_val = ((tax[i]/tax[i-1])-1)*100      if tax[i-1]   else 0.0
        s_scores.append(score_param(s_val)); p_scores.append(score_param(p_val))
        o_scores.append(score_param(o_val)); r_scores.append(score_param(r_val))
        t_scores.append(score_param(t_val))
        year_labels.append(years[i-1] if i-1 < len(years) else "Yr"+str(i))

    s_total = sum(s_scores); p_total = sum(p_scores); o_total = sum(o_scores)
    r_total = sum(r_scores); t_total = sum(t_scores)
    total   = s_total + p_total + o_total + r_total + t_total

    if not mc_pass:     verdict = "REJECTED — Wrong Market Cap"
    elif de_pass is False: verdict = "REJECTED — High Debt (D/E>=0.3)"
    elif total > MDSPORT_VALID: verdict = "VALID"
    elif total >= MDSPORT_WATCH: verdict = "WATCHLIST"
    else:               verdict = "REJECTED — Low Score"
    if is_fin:          verdict += " (informational — Bank/NBFC)"

    return {
        "symbol": ticker, "company_name": name, "market_cap_cr": mc,
        "is_financial": is_fin, "de_ratio": round(de_ratio,3) if de_ratio is not None else None,
        "de_pass": de_pass, "mc_pass": mc_pass, "passes_gates": passes_gates,
        "total_score": total, "verdict": verdict,
        "s_score": s_total, "p_score": p_total, "o_score": o_total,
        "r_score": r_total, "t_score": t_total,
        "years_used": n-1,
        "years_range": f"{year_labels[0]} – {year_labels[-1]}" if year_labels else "",
        "status": "OK", "error": ""
    }


def check_mptds(symbol, session, level="moderate", years="max"):
    url = f"{BASE}/api/mptds/{symbol}?level={level}&years={years}"
    try:
        r = session.get(url, timeout=REQUEST_TIMEOUT)
        if r.status_code == 404:
            return _mptds_err(symbol, "NOT_FOUND", "Symbol not found")
        if r.status_code == 429:
            time.sleep(6); r = session.get(url, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
        return _mptds_parse(symbol, r.json())
    except requests.exceptions.Timeout: return _mptds_err(symbol, "TIMEOUT", "Timed out")
    except requests.exceptions.ConnectionError: return _mptds_err(symbol, "CONN_ERROR", "Connection failed")
    except Exception as e: return _mptds_err(symbol, "ERROR", str(e))


def _mptds_err(symbol, status, msg):
    return {"symbol": symbol, "company_name": "", "market_cap_cr": None,
            "is_large_cap": False, "verdict": status, "mptds_valid": False,
            "p_trend":"","t_trend":"","d_trend":"","s_trend":"",
            "data_points": None, "status": status, "error": msg}


def _mptds_parse(symbol, data):
    tr = data.get("trends", {})
    def tp(k): p = tr.get(k,{}); return p.get("trend",""), p.get("slope"), p.get("r2"), p.get("count")
    pt,ps,pr,pc = tp("P"); tt,ts,tr2,_ = tp("T"); dt,ds,dr,_ = tp("D"); st2,ss,sr,_ = tp("S")
    mc      = float(data.get("marketCap") or 0)
    verdict = str(data.get("verdict",""))
    valid   = verdict.upper() == "VALID" and mc >= MPTDS_MC_MIN
    return {
        "symbol": data.get("ticker", symbol), "company_name": data.get("companyName",""),
        "market_cap_cr": mc or None, "is_large_cap": mc >= MPTDS_MC_MIN,
        "verdict": verdict, "mptds_valid": valid,
        "p_trend": pt, "p_slope_pct": ps, "p_r2": pr,
        "t_trend": tt, "t_slope_pct": ts, "t_r2": tr2,
        "d_trend": dt, "d_slope_pct": ds, "d_r2": dr,
        "s_trend": st2, "s_slope_pct": ss, "s_r2": sr,
        "data_points": pc, "status": "OK", "error": ""
    }


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(df_mds, df_mpt):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if df_mds is not None:
            valid_mds = df_mds[df_mds["verdict"]=="VALID"]
            valid_mds.to_excel(writer, sheet_name="MDSPORT Valid", index=False)
            df_mds.to_excel(writer,    sheet_name="MDSPORT All",   index=False)
        if df_mpt is not None:
            valid_mpt = df_mpt[df_mpt["mptds_valid"]==True]
            valid_mpt.to_excel(writer, sheet_name="MPTDS Valid", index=False)
            df_mpt.to_excel(writer,    sheet_name="MPTDS All",   index=False)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        buf.seek(0)
        wb = load_workbook(buf)
        H_FILL = PatternFill("solid", fgColor="1F3864")
        H_FONT = Font(color="FFFFFF", bold=True)
        G_FILL = PatternFill("solid", fgColor="C6EFCE")
        Y_FILL = PatternFill("solid", fgColor="FFEB9C")
        R_FILL = PatternFill("solid", fgColor="FFC7CE")
        for sname in wb.sheetnames:
            ws = wb[sname]
            for cell in ws[1]:
                cell.fill = H_FILL; cell.font = H_FONT
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.row_dimensions[1].height = 28
            ws.freeze_panes = "A2"
            for col in ws.columns:
                w = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+3, 38)
        for sname in ["MDSPORT All", "MPTDS All"]:
            if sname not in wb.sheetnames: continue
            ws = wb[sname]
            vcol = next((i for i,c in enumerate(ws[1],1) if "verdict" in str(c.value or "").lower()), None)
            if vcol:
                for row in ws.iter_rows(min_row=2):
                    v = str(row[vcol-1].value or "").upper()
                    fill = G_FILL if "VALID" in v and "REJECTED" not in v else (Y_FILL if "WATCH" in v else R_FILL)
                    for cell in row: cell.fill = fill
        buf2 = io.BytesIO()
        wb.save(buf2)
        return buf2.getvalue()
    except:
        buf.seek(0)
        return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  UI HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def verdict_badge(verdict):
    v = str(verdict).upper()
    if "VALID" in v and "REJECTED" not in v:
        return f'<span class="valid-badge">✅ VALID</span>'
    elif "WATCH" in v:
        return f'<span class="watch-badge">👁 WATCHLIST</span>'
    else:
        return f'<span class="reject-badge">❌ {verdict}</span>'

def trend_arrow(t):
    if t == "increasing":  return "↑"
    if t == "decreasing":  return "↓"
    return "~"

def fmt_cr(val):
    if val is None or pd.isna(val): return "N/A"
    return f"₹{val:,.0f} Cr"


# ══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## ⚙️ Settings")
    st.divider()

    mode = st.radio(
        "Screener mode",
        ["Both MDSPORT + MPTDS", "MDSPORT only (small cap)", "MPTDS only (large cap)"],
        index=0
    )

    st.divider()

    universe = st.radio(
        "Stock universe",
        ["Quick test — 50 stocks (~2 min)", "Nifty 500 (~12 min)", "Full NSE — 2,258 stocks (~90 min)"],
        index=0
    )

    custom_input = st.text_input(
        "Or enter specific symbols (comma separated)",
        placeholder="e.g. RELIANCE, TCS, DIXON",
        help="Leave blank to use the universe selection above"
    )

    st.divider()
    st.markdown("**MPTDS settings**")
    level = st.select_slider("Sensitivity", ["moderate", "strict"], value="moderate")
    years = st.select_slider("Data period", ["max", "5yr"], value="max")

    st.divider()
    delay = st.slider("Delay between requests (sec)", 1.0, 3.0, 1.5, 0.1,
                       help="Higher = slower but less risk of rate limiting")

    st.divider()
    col1, col2 = st.columns(2)
    with col1:
        run_btn = st.button("▶ Run", type="primary", use_container_width=True)
    with col2:
        reset_btn = st.button("🗑 Reset", use_container_width=True)

    if reset_btn:
        st.session_state.mds_results = {}
        st.session_state.mpt_results = {}
        st.session_state.log = []
        st.session_state.done = 0
        if os.path.exists(CHECKPOINT_FILE):
            os.remove(CHECKPOINT_FILE)
        st.success("Reset done!")

    st.divider()
    st.markdown("""
    <small style='color:#888'>
    Data source: nesliferesearch.com<br>
    (Screener.in + Yahoo Finance)<br><br>
    <b>MDSPORT</b>: Small cap ₹1k–10k Cr, D/E &lt; 0.3<br>
    <b>MPTDS</b>: Large cap ≥ ₹10k Cr, all trends ↑
    </small>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN PAGE
# ══════════════════════════════════════════════════════════════════════════════

st.markdown('<p class="main-title">📈 NSE Stock Screener</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">MDSPORT + MPTDS · Wealthcon Framework · All 2,258 NSE listed stocks</p>', unsafe_allow_html=True)
st.divider()

# ── Metric cards ──────────────────────────────────────────────────────────────
mds_all   = list(st.session_state.mds_results.values())
mpt_all   = list(st.session_state.mpt_results.values())
mds_valid = [r for r in mds_all if r.get("verdict") == "VALID"]
mpt_valid = [r for r in mpt_all if r.get("mptds_valid")]

c1, c2, c3, c4, c5 = st.columns(5)
with c1:
    st.metric("Screened",       len(mds_all) + len(mpt_all))
with c2:
    st.metric("MDSPORT Valid",  len(mds_valid), delta=None)
with c3:
    st.metric("MPTDS Valid",    len(mpt_valid), delta=None)
with c4:
    mds_watch = len([r for r in mds_all if "WATCHLIST" in str(r.get("verdict",""))])
    st.metric("MDSPORT Watchlist", mds_watch)
with c5:
    errs = len([r for r in mds_all+mpt_all if r.get("status") not in ("OK","NOT_FOUND","INSUFFICIENT_DATA")])
    st.metric("Errors", errs)

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
#  RUN SCREENER
# ══════════════════════════════════════════════════════════════════════════════

if run_btn:
    # Determine symbols
    if custom_input.strip():
        symbols = [s.strip().upper() for s in custom_input.split(",") if s.strip()]
    elif "50 stocks" in universe:
        symbols = None; limit = 50
    elif "Nifty 500" in universe:
        with st.spinner("Fetching Nifty 500 symbols..."):
            symbols = fetch_nse_symbols()[:500]
        limit = None
    else:
        with st.spinner("Fetching all NSE symbols..."):
            symbols = fetch_nse_symbols()
        limit = None

    if symbols is None:
        with st.spinner("Fetching NSE symbols..."):
            symbols = fetch_nse_symbols()
        if "50 stocks" in universe:
            symbols = symbols[:50]

    # Load checkpoint
    cp = {"mdsport": {}, "mptds": {}}
    if os.path.exists(CHECKPOINT_FILE):
        try:
            with open(CHECKPOINT_FILE) as f:
                cp = json.load(f)
            st.info(f"♻️ Resuming from checkpoint: {len(cp.get('mdsport',{}))} MDSPORT + {len(cp.get('mptds',{}))} MPTDS already done")
        except:
            pass

    mds_done = dict(cp.get("mdsport", {}))
    mpt_done = dict(cp.get("mptds",   {}))

    do_mds = "MPTDS only" not in mode
    do_mpt = "MDSPORT only" not in mode

    mds_todo = [s for s in symbols if s not in mds_done] if do_mds else []
    mpt_todo = [s for s in symbols if s not in mpt_done] if do_mpt else []
    total_todo = len(mds_todo) + len(mpt_todo)

    session = requests.Session()
    session.headers.update(HEADERS)

    # ── Progress UI ──────────────────────────────────────────────────────────
    progress_bar  = st.progress(0.0)
    status_text   = st.empty()
    live_log      = st.empty()
    done_count    = len(mds_done) + len(mpt_done)
    total_count   = len(symbols) * (2 if "Both" in mode else 1)
    log_lines     = []

    # ── MDSPORT loop ─────────────────────────────────────────────────────────
    if do_mds and mds_todo:
        status_text.markdown(f"**MDSPORT scan** — {len(mds_todo)} stocks remaining...")
        for i, sym in enumerate(mds_todo):
            raw = fetch_company_data(sym, session)
            row = compute_mdsport(sym, raw)
            mds_done[sym] = row

            v = str(row.get("verdict",""))
            if v == "VALID":
                mc  = fmt_cr(row.get("market_cap_cr"))
                de  = f"D/E={row['de_ratio']:.2f}" if row.get("de_ratio") is not None else ""
                scr = row.get("total_score","")
                log_lines.insert(0, f"✅ **{sym}** — {row.get('company_name','')} | {mc} | {de} | Score {scr}/100")
            elif "WATCH" in v:
                log_lines.insert(0, f"👁 **{sym}** — WATCHLIST | Score {row.get('total_score','')}/100")

            if len(log_lines) > 15:
                log_lines = log_lines[:15]

            done_count += 1
            progress_bar.progress(min(done_count / max(total_count,1), 1.0))
            status_text.markdown(f"**MDSPORT** — {i+1}/{len(mds_todo)} | ✅ Valid so far: {len([r for r in mds_done.values() if r.get('verdict')=='VALID'])}")
            live_log.markdown("\n".join(log_lines) if log_lines else "*No VALID stocks yet...*")

            if (i+1) % 25 == 0:
                cp["mdsport"] = mds_done
                with open(CHECKPOINT_FILE,"w") as f: json.dump(cp, f)

            if i < len(mds_todo)-1:
                time.sleep(delay)

        cp["mdsport"] = mds_done
        with open(CHECKPOINT_FILE,"w") as f: json.dump(cp, f)

    # ── MPTDS loop ────────────────────────────────────────────────────────────
    if do_mpt and mpt_todo:
        status_text.markdown(f"**MPTDS scan** — {len(mpt_todo)} stocks remaining...")
        for i, sym in enumerate(mpt_todo):
            row = check_mptds(sym, session, level=level, years=years)
            mpt_done[sym] = row

            if row.get("mptds_valid"):
                mc = fmt_cr(row.get("market_cap_cr"))
                log_lines.insert(0, f"🏆 **{sym}** (MPTDS) — {row.get('company_name','')} | {mc} | P{trend_arrow(row.get('p_trend',''))} T{trend_arrow(row.get('t_trend',''))} D{trend_arrow(row.get('d_trend',''))} S{trend_arrow(row.get('s_trend',''))}")

            if len(log_lines) > 15:
                log_lines = log_lines[:15]

            done_count += 1
            progress_bar.progress(min(done_count / max(total_count,1), 1.0))
            status_text.markdown(f"**MPTDS** — {i+1}/{len(mpt_todo)} | 🏆 Valid so far: {len([r for r in mpt_done.values() if r.get('mptds_valid')])}")
            live_log.markdown("\n".join(log_lines) if log_lines else "*Scanning...*")

            if (i+1) % 25 == 0:
                cp["mptds"] = mpt_done
                with open(CHECKPOINT_FILE,"w") as f: json.dump(cp, f)

            if i < len(mpt_todo)-1:
                time.sleep(delay)

        cp["mptds"] = mpt_done
        with open(CHECKPOINT_FILE,"w") as f: json.dump(cp, f)

    # ── Save to session state ─────────────────────────────────────────────────
    st.session_state.mds_results = mds_done
    st.session_state.mpt_results = mpt_done
    progress_bar.progress(1.0)
    status_text.markdown("✅ **Screening complete!**")
    st.balloons()
    st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
#  RESULTS TABLES
# ══════════════════════════════════════════════════════════════════════════════

tab1, tab2, tab3, tab4 = st.tabs([
    "✅ MDSPORT Valid",
    "📊 MDSPORT All Results",
    "🏆 MPTDS Valid",
    "📋 MPTDS All Results"
])

# ── Tab 1: MDSPORT Valid ──────────────────────────────────────────────────────
with tab1:
    if mds_valid:
        df = pd.DataFrame(mds_valid).sort_values("total_score", ascending=False)
        st.markdown(f"### {len(mds_valid)} MDSPORT Valid Companies")
        st.markdown("*Passed M+D gates (Market Cap ₹1k–10k Cr, D/E < 0.3) and scored > 75/100*")

        display_cols = ["symbol","company_name","market_cap_cr","de_ratio",
                        "total_score","s_score","p_score","o_score","r_score","t_score","years_range"]
        rename = {"symbol":"Symbol","company_name":"Company","market_cap_cr":"Mkt Cap (Cr)",
                  "de_ratio":"D/E","total_score":"Score","s_score":"S","p_score":"P",
                  "o_score":"O","r_score":"R","t_score":"T","years_range":"Period"}
        df_show = df[[c for c in display_cols if c in df.columns]].rename(columns=rename)
        st.dataframe(df_show, use_container_width=True, hide_index=True)
    else:
        st.info("No MDSPORT results yet. Run the screener using the sidebar.")

# ── Tab 2: MDSPORT All ───────────────────────────────────────────────────────
with tab2:
    if mds_all:
        df_all = pd.DataFrame(mds_all)
        st.markdown(f"### All {len(mds_all)} MDSPORT Results")

        # Filter controls
        col1, col2 = st.columns(2)
        with col1:
            verdict_filter = st.multiselect(
                "Filter by verdict",
                options=df_all["verdict"].unique().tolist() if "verdict" in df_all.columns else [],
                default=[]
            )
        with col2:
            search = st.text_input("Search symbol or company", "")

        df_show = df_all.copy()
        if verdict_filter:
            df_show = df_show[df_show["verdict"].isin(verdict_filter)]
        if search:
            mask = (df_show["symbol"].str.contains(search.upper(), na=False) |
                    df_show["company_name"].str.contains(search, case=False, na=False))
            df_show = df_show[mask]

        show_cols = ["symbol","company_name","market_cap_cr","de_ratio","mc_pass","de_pass","total_score","verdict","s_score","p_score","o_score","r_score","t_score"]
        df_show = df_show[[c for c in show_cols if c in df_show.columns]]
        st.dataframe(df_show, use_container_width=True, hide_index=True)
    else:
        st.info("No MDSPORT results yet.")

# ── Tab 3: MPTDS Valid ────────────────────────────────────────────────────────
with tab3:
    if mpt_valid:
        df = pd.DataFrame(mpt_valid).sort_values("market_cap_cr", ascending=False)
        st.markdown(f"### {len(mpt_valid)} MPTDS Valid Companies")
        st.markdown("*Market Cap ≥ ₹10,000 Cr with all 4 trends increasing (Price ↑, Tax ↑, Dividend ↑, Sales ↑)*")

        df["Trends"] = (df["p_trend"].apply(trend_arrow) + " " +
                        df["t_trend"].apply(trend_arrow) + " " +
                        df["d_trend"].apply(trend_arrow) + " " +
                        df["s_trend"].apply(trend_arrow))
        show_cols = ["symbol","company_name","market_cap_cr","Trends","p_slope_pct","t_slope_pct","d_slope_pct","s_slope_pct","data_points"]
        rename = {"symbol":"Symbol","company_name":"Company","market_cap_cr":"Mkt Cap (Cr)",
                  "p_slope_pct":"P Growth%","t_slope_pct":"T Growth%",
                  "d_slope_pct":"D Growth%","s_slope_pct":"S Growth%","data_points":"Years"}
        df_show = df[[c for c in show_cols if c in df.columns]].rename(columns=rename)
        st.dataframe(df_show, use_container_width=True, hide_index=True)
    else:
        st.info("No MPTDS results yet. Run the screener using the sidebar.")

# ── Tab 4: MPTDS All ─────────────────────────────────────────────────────────
with tab4:
    if mpt_all:
        df_all = pd.DataFrame(mpt_all)
        st.markdown(f"### All {len(mpt_all)} MPTDS Results")

        col1, col2 = st.columns(2)
        with col1:
            v_filter = st.multiselect(
                "Filter by verdict",
                options=df_all["verdict"].unique().tolist() if "verdict" in df_all.columns else [],
                default=[], key="mptds_filter"
            )
        with col2:
            search2 = st.text_input("Search symbol or company", "", key="mptds_search")

        df_show = df_all.copy()
        if v_filter: df_show = df_show[df_show["verdict"].isin(v_filter)]
        if search2:
            mask = (df_show["symbol"].str.contains(search2.upper(), na=False) |
                    df_show["company_name"].str.contains(search2, case=False, na=False))
            df_show = df_show[mask]

        show_cols = ["symbol","company_name","market_cap_cr","verdict","p_trend","t_trend","d_trend","s_trend","data_points"]
        df_show = df_show[[c for c in show_cols if c in df_show.columns]]
        st.dataframe(df_show, use_container_width=True, hide_index=True)
    else:
        st.info("No MPTDS results yet.")


# ══════════════════════════════════════════════════════════════════════════════
#  DOWNLOAD SECTION
# ══════════════════════════════════════════════════════════════════════════════

st.divider()
st.markdown("### 📥 Download Results")
col1, col2, col3 = st.columns(3)

df_mds = pd.DataFrame(mds_all) if mds_all else None
df_mpt = pd.DataFrame(mpt_all) if mpt_all else None

with col1:
    if mds_valid:
        csv = pd.DataFrame(mds_valid).to_csv(index=False).encode("utf-8")
        st.download_button("⬇ MDSPORT Valid (CSV)", csv,
                           file_name="nse_mdsport_valid.csv", mime="text/csv",
                           use_container_width=True)
    else:
        st.button("⬇ MDSPORT Valid (CSV)", disabled=True, use_container_width=True)

with col2:
    if mpt_valid:
        csv2 = pd.DataFrame(mpt_valid).to_csv(index=False).encode("utf-8")
        st.download_button("⬇ MPTDS Valid (CSV)", csv2,
                           file_name="nse_mptds_valid.csv", mime="text/csv",
                           use_container_width=True)
    else:
        st.button("⬇ MPTDS Valid (CSV)", disabled=True, use_container_width=True)

with col3:
    if mds_all or mpt_all:
        excel_bytes = build_excel(df_mds, df_mpt)
        st.download_button("⬇ Full Results (Excel)", excel_bytes,
                           file_name=f"nse_results_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
    else:
        st.button("⬇ Full Results (Excel)", disabled=True, use_container_width=True)

st.divider()
st.caption("NSE Screener · Wealthcon Framework · Data from nesliferesearch.com (Screener.in + Yahoo Finance)")
